import io
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Blueprint, request, render_template, redirect, url_for, abort, send_file
from app.database import get_db_connection
from app.engine import calculate_amortization_engine

# Explicit Blueprint registration matching template endpoints
main = Blueprint('main', __name__)
bp = main  # Compatibility alias hook for application factories

# ==========================================
# 1. ROUTE: MASTER PORTFOLIO EDITOR
# ==========================================
@main.route('/')
@main.route('/portfolio')
def portfolio_overview():
    """
    Renders the broad master database grid layout.
    """
    with get_db_connection() as conn:
        loans_raw = conn.execute("SELECT * FROM tbl_amortization_detail").fetchall()
        
    portfolio_loans = []
    for loan in loans_raw:
        loan_dict = dict(loan)
        loan_dict['term_months'] = int(loan_dict.get('term_years', 0)) * 12
        loan_dict['accrued'] = float(loan_dict.get('total_interest', 0.0))
        loan_dict['paid'] = float(loan_dict.get('total_paid', 0.0))
        portfolio_loans.append(loan_dict)

    return render_template('portfolio_editor.html', portfolio_loans=portfolio_loans)


# ==========================================
# 2. ROUTE: BATCH SHEET SAVE OPERATION
# ==========================================
@main.route('/portfolio/save', methods=['POST'])
def save_portfolio_changes():
    """
    Processes inline cell matrix modifications committed via the browser spreadsheet form.
    """
    try:
        loan_ids = request.form.getlist('loan_id[]')
        principals = request.form.getlist('principal[]')
        rates = request.form.getlist('rate[]')
        term_years_list = request.form.getlist('term_years[]')
        balloon_years_list = request.form.getlist('balloon_years[]')
        start_dates = request.form.getlist('start_date[]')

        with get_db_connection() as conn:
            cursor = conn.cursor()
            for idx in range(len(loan_ids)):
                l_id = loan_ids[idx].strip()
                p_val = float(principals[idx])
                r_val = float(rates[idx])
                t_yrs = int(term_years_list[idx])
                b_yrs = int(balloon_years_list[idx] or 0)
                s_date = start_dates[idx].strip()

                cursor.execute("""
                    UPDATE tbl_amortization_detail 
                    SET principal = ?, rate = ?, term_years = ?, balloon_years = ?, start_date = ?
                    WHERE loan_id = ?
                """, (p_val, r_val, t_yrs, b_yrs, s_date, l_id))
            conn.commit()

        for loan_id in loan_ids:
            calculate_amortization_engine(loan_id.strip())

        return redirect(url_for('main.portfolio_overview'))
    except Exception as e:
        return f"Batch Execution Fallback Triggered. Error: {str(e)}", 500


# ==========================================
# 3. ROUTE: LIVE DATABASE DATA EXCEL EXPORT
# ==========================================
@main.route('/portfolio/export')
def export_portfolio_excel():
    """
    Extracts current database status logs and packages them into an executive spreadsheet sheet stream.
    """
    try:
        with get_db_connection() as conn:
            loans = conn.execute("SELECT * FROM tbl_amortization_detail").fetchall()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Active Portfolio"
        ws.views.sheetView[0].showGridLines = True
        
        # Styles Layout Definition Frame
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_body = Font(name="Segoe UI", size=11, color="000000")
        font_bold = Font(name="Segoe UI", size=11, bold=True, color="000000")
        
        fill_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_summary = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        
        border_thin = Side(border_style="thin", color="E2E8F0")
        border_double = Side(border_style="double", color="334155")
        
        cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
        summary_border = Border(top=border_thin, bottom=border_double)
        
        headers = [
            "Loan ID", "Loan Name", "Start Date", "Principal", 
            "Rate %", "Tenure (Year)", "Balloon Period (Year)", "Comments"
        ]
        
        ws.append(headers)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center if "Date" in header or "Year" in header or "ID" in header else (align_right if "Principal" in header or "Rate" in header else align_left)
        
        for row_idx, loan in enumerate(loans, 2):
            loan_dict = dict(loan)
            row_data = [
                loan_dict.get('loan_id', ''),
                loan_dict.get('loan_name', ''),
                loan_dict.get('start_date', ''),
                float(loan_dict.get('principal', 0.0)),
                float(loan_dict.get('rate', 0.0)),
                int(loan_dict.get('term_years', 0)),
                int(loan_dict.get('balloon_years', 0)),
                loan_dict.get('comments', '') or "No modifications logged."
            ]
            ws.append(row_data)
            is_even = (row_idx % 2 == 0)
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_body
                if is_even:
                    cell.fill = fill_zebra
                cell.border = cell_border
                
                if col_idx in [1, 3, 6, 7]:
                    cell.alignment = align_center
                elif col_idx == 4:
                    cell.number_format = "$#,##0.00"
                    cell.alignment = align_right
                elif col_idx == 5:
                    cell.number_format = "0.00"
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left
        
        # Summary aggregation calculation line
        summary_row_idx = len(loans) + 2
        ws.cell(row=summary_row_idx, column=1, value="Total Summary Portfolio").font = font_bold
        ws.cell(row=summary_row_idx, column=1).alignment = align_left
        ws.cell(row=summary_row_idx, column=4, value=f"=SUM(D2:D{summary_row_idx-1})").font = font_bold
        ws.cell(row=summary_row_idx, column=4).number_format = "$#,##0.00"
        ws.cell(row=summary_row_idx, column=4).alignment = align_right
        
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=summary_row_idx, column=col_idx)
            c.border = summary_border
            c.fill = fill_summary
            
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
            
        ws.freeze_panes = "A2"
        
        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        
        return send_file(
            excel_stream,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="Portfolio_Active_Export.xlsx"
        )
    except Exception as e:
        return f"Database Data Export Interruption: {str(e)}", 500


# ==========================================
# 4. NEW ROUTE: DOWNLOAD BLANK TEMPLATE FILE
# ==========================================
@main.route('/portfolio/template')
def download_blank_template():
    """
    Generates and downloads a correctly structured blank template layout frame.
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Upload Template"
        ws.views.sheetView[0].showGridLines = True
        
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="334155", end_color="334155", fill_type="solid") # Dark Slate Gray
        align_center = Alignment(horizontal="center", vertical="center")
        
        headers = [
            "Loan ID", "Loan Name", "Start Date", "Principal", 
            "Rate %", "Tenure (Year)", "Balloon Period (Year)", "Comments"
        ]
        
        ws.append(headers)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = 18
            
        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        
        return send_file(
            excel_stream,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="Portfolio_Upload_Template.xlsx"
        )
    except Exception as e:
        return f"Template Configuration Export Interruption: {str(e)}", 500


# ==========================================
# 5. UPDATED ROUTE: MASS EXCEL UPLOAD PARSER
# ==========================================
@main.route('/api/loan/upload', methods=['POST'])
def handle_bulk_upload():
    """
    Parses complex multi-column binary data structured directly out of spreadsheet files.
    """
    if 'file' not in request.files:
        return "No configuration file uploaded", 400
    file = request.files['file']
    if file.filename == '' or not file.filename.endswith(('.xlsx', '.xlsm')):
        return "Unsupported file extension pattern. Upload standard openxml sheets (.xlsx)", 400

    try:
        # Load the uploaded raw excel file binary stream array directly into openpyxl
        file_stream = io.BytesIO(file.read())
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        ws = wb.active  # Target active ledger data sheet tab
        
        processed_loan_ids = []

        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Row index starts at 2 to skip headers column indices row
            for r_idx in range(2, ws.max_row + 1):
                # Safely pull values matching the exact layout column indexing offsets
                l_id = str(ws.cell(row=r_idx, column=1).value or '').strip()
                name = str(ws.cell(row=r_idx, column=2).value or '').strip()
                s_date_val = ws.cell(row=r_idx, column=3).value
                p_val = ws.cell(row=r_idx, column=4).value
                r_val = ws.cell(row=r_idx, column=5).value
                t_yrs = ws.cell(row=r_idx, column=6).value
                b_yrs = ws.cell(row=r_idx, column=7).value
                comments = str(ws.cell(row=r_idx, column=8).value or '').strip()

                # Core constraint protection validation checks
                if not l_id or not name:
                    continue  # Break early if critical identifiers row are missing data
                
                # Sanitize calculations constraints numerical data elements safely
                p_val = float(p_val) if p_val is not None else 0.0
                r_val = float(r_val) if r_val is not None else 0.0
                t_yrs = int(t_yrs) if t_yrs is not None else 0
                b_yrs = int(b_yrs) if b_yrs is not None else 0
                
                # Date structural standard format clean conversion handling
                if isinstance(s_date_val, datetime):
                    s_date = s_date_val.strftime('%Y-%m-%d')
                else:
                    s_date = str(s_date_val or '').strip()

                # Clean tracking database entries for targeted loan paths
                cursor.execute("DELETE FROM tbl_amortization_calc WHERE loan_id = ?", (l_id,))
                cursor.execute("DELETE FROM tbl_amortization_detail WHERE loan_id = ?", (l_id,))
                
                cursor.execute("""
                    INSERT INTO tbl_amortization_detail 
                    (loan_id, loan_name, principal, rate, term_years, balloon_years, start_date, total_interest, total_paid, comments)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 0, 0, ?)
                """, (l_id, name, p_val, r_val, t_yrs, b_yrs, s_date, comments))
                processed_loan_ids.append(l_id)
                
            conn.commit()

        # Run amortization schedule engine calculations logic 
        for loan_id in processed_loan_ids:
            calculate_amortization_engine(loan_id)

        return redirect(url_for('main.portfolio_overview'))
    except Exception as e:
        return f"Excel Binary Upload Parsing Exception: {str(e)}", 500


# ==========================================
# 6. ROUTE: GRANULAR AMORTIZATION ENGINE VIEW
# ==========================================
@main.route('/loan/<loan_id>')
@main.route('/loan', defaults={'loan_id': None})
def loan_detail_view(loan_id):
    with get_db_connection() as conn:
        loans = conn.execute("SELECT * FROM tbl_amortization_detail").fetchall()
        selected_loan = None
        schedule = []
        
        if loan_id:
            selected_loan = conn.execute("SELECT * FROM tbl_amortization_detail WHERE loan_id = ?", (loan_id,)).fetchone()
            if not selected_loan:
                abort(404)
            schedule = conn.execute("SELECT * FROM tbl_amortization_calc WHERE loan_id = ? ORDER BY month_index ASC", (loan_id,)).fetchall()
            
    return render_template('amortization_ledger.html', loans=loans, selected_loan=selected_loan, schedule=schedule)


# ==========================================
# 7. ROUTE: SINGLE MANUAL ASSET PROVISION
# ==========================================
@main.route('/api/loan/create', methods=['POST'])
def handle_create_loan():
    try:
        loan_id = request.form['loan_id'].strip()
        loan_name = request.form['loan_name'].strip()
        principal = float(request.form['principal'])
        rate = float(request.form['rate'])
        term_years = int(request.form['term_years'])
        balloon_years = int(request.form['balloon_years'] or 0)
        start_date = request.form['start_date'].strip()

        with get_db_connection() as conn:
            existing = conn.execute("SELECT 1 FROM tbl_amortization_detail WHERE loan_id = ?", (loan_id,)).fetchone()
            if existing:
                return f"Constraint Error: Identifier '{loan_id}' already tracking.", 400

            conn.execute("""
                INSERT INTO tbl_amortization_detail (loan_id, loan_name, principal, rate, term_years, balloon_years, start_date, total_interest, total_paid, comments)
                VALUES (?, ?, ?, ?, ?, ?, ?, 0, 0, 'Created manually via portal panel.')
            """, (loan_id, loan_name, principal, rate, term_years, balloon_years, start_date))
            conn.commit()

        calculate_amortization_engine(loan_id)
        return redirect(url_for('main.portfolio_overview'))
    except Exception as e:
        return f"Operational Processing Failure: {str(e)}", 500